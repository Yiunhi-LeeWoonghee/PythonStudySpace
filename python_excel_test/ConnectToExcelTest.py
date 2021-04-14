# openpyxlライブラリが必要
# ターミナルにて pip install openpyxl 実行
# コマンドライン引数取得
import sys
# no moduleエラー解消（ライブラリ格納先を明視）
sys.path.append('c:\\users\\thele\\appdata\\local\\programs\\python\\python37\\lib\\site-packages')

# pythonとexcelの連携ライブラリインポート
import openpyxl
# 複数ファイルを処理するため、globライブラリをインポート
import glob

# コマンドライン引数一覧
# args[0] = python実行ファイル名
# args[1] = excelファイル格納先（ディレクトリ）
# args[2] = 結果物出力先（ディレクトリ）

# 引数を受け取る
# args = sys.argv
# execFileName = args[0]
# excelDir = args[1]
# resultDir = args[2]

execFileName = "ConnectToExcelTest.py"
excelDir = "D:\\devStudy\\PythonStudySpace\\excel_test"
resultDir = "D:\\devStudy\\PythonStudySpace\\excel_tes\\result"
excelExtn = "\\*.xlsx"

# ディレクトリ及びExcelファイル指定
files = glob.glob(excelDir + excelExtn)

sheet_index = 0
# Excelファイル読み込み
for excelfile in files:
    workbook = openpyxl.load_workbook(excelfile)
    # シート取得
    for sheetName in [workbook.sheetnames]:
        workSheet = workbook[sheet_index]
        m_row = workSheet.max_row
        m_col = workSheet.max_column
        # 各シートのセル取得
        for i in m_row:
            workSheet.cell(column=m_col-i, row=i)